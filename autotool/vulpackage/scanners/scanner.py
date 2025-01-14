import re
from pprint import pprint
from vulpackage.core.consolidation import Consolidation
from openpyxl import load_workbook

from colorama import init
from termcolor import colored
import textwrap
init()
from terminaltables import SingleTable, DoubleTable
import subprocess

from dotenv import load_dotenv, find_dotenv
import os
import paramiko

load_dotenv(find_dotenv())
import xlsxwriter


class Scanner:

    def __init__(self):

        self.consolidation = Consolidation()





    def scan(self):
        pass

    def get_scan_status(self):
        pass

    def get_scan_results(self):
        pass

    def is_valid_scan(self):
        pass

    def list_scans(self):
        pass

    def pause(self):
        pass

    def resume(self):
        pass
    def start_sp(self):
        pass

    def stop(self):
        pass

    def remove(self):
        pass

    def is_duplicate(self):
        pass

    def _get_address(self, target):
        ex = '(?:http.*://)?(?P<host>[^:/ ]+).?(?P<port>[0-9]*).*'
        h = re.search(ex, target)
        return h.group('host')

    def SSH(self, u, h, c, p):
        try:
            self.ssh = paramiko.SSHClient()
            self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            self.ssh.connect(hostname=h, username=u, password=p)
            stdin, stdout, stderr = self.ssh.exec_command(c)
            print(''.join(stdout.readlines()))

        except paramiko.AuthenticationException:
            print("Authentication failed when connecting to %s" % h)
        except:
            print("Could not SSH to %s" % h)
            self.ssh.close()

    def disable_scanner_services(self, config):
        if config['disable_serviceslw']:
            process = subprocess.Popen(r'openvas-stop&"C:\Program Files\rapid7\nexpose\nsc\nscsvc.bat" stop',
                                       stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True,
                                       universal_newlines=True)
            while True:
                output = process.stdout.readline()
                print(output.strip())
                return_code = process.poll()
                if return_code is not None:
                    break
            
        if config['disable_servicesrw']:
            
            self.SSH(os.getenv('R_USERNAME'), os.getenv('R_HOST'),
                     r'openvas-stop&"C:\Program Files\rapid7\nexpose\nsc\nscsvc.bat" stop', os.getenv('R_PASSWORD'))
        if config['disable_servicesll']:
            
            RCmd = os.popen('openvas-stop;service nexposeconsole stop;systemctl stop nexposeconsole')
            print(RCmd)
        if config['disable_servicesrl']:
            
            self.SSH(os.getenv('R_USERNAME'), os.getenv('R_HOST'),
                     'openvas-stop;service nexposeconsole stop;systemctl stop nexposeconsole', os.getenv('R_PASSWORD'))

    def enable_scanner_services(self, config):
        if config['enable_serviceslw']:
            RCmd = os.popen(r'openvas-start&"C:\Program Files\rapid7\nexpose\nsc\nscsvc.bat" start').read()
            print(RCmd)
        if config['enable_servicesrw']:
            self.SSH(os.getenv('R_USERNAME'), os.getenv('R_HOST'),
                     r'openvas-start&"C:\Program Files\rapid7\nexpose\nsc\nscsvc.bat" start ', os.getenv('R_PASSWORD'))
        if config['enable_servicesll']:
            RCmd = os.popen('/usr/share/zaproxy/zap.sh -session /root/Desktop/zap/suba.session;openvas-start;service nexposeconsole start;systemctl start nexposeconsole').read()
            print(RCmd)
        if config['enable_servicesrl']:
            self.SSH(os.getenv('R_USERNAME'), os.getenv('R_HOST'),
                     '/usr/share/zaproxy/zap.sh -session /root/Desktop/zap/suba.session;openvas-start;service nexposeconsole start;systemctl start nexposeconsole',
                     os.getenv('R_PASSWORD'))

    def print_scan_status(self, scan_status_list):
        status = []
        status.append(['#', 'Scanner', 'Status'])
        count = 0

        for scan_status in scan_status_list:
            count += 1
            status.append([count, scan_status['scanner'], scan_status['status']])

        status_table = DoubleTable(status)
        status_table.title = 'Scan Status'
        print(status_table.table)

    def print_report(self, zscan_results, nscan_results, oscan_results):

        # if not zscan_results and nscan_results and oscan_results :

        # return False

        results1 = list(zscan_results.values())
        results2 = list(nscan_results.values())
        results3 = list(oscan_results.values())
        allresults = [results1, results2, results3]

        zscan_report = []
        nscan_report = []
        oscan_report = []

        final = [['#', 'Vulnerability Name', 'Risk', 'Severity', 'CVE/CWE ID','URLs','Scanner']]
        dup=[['#', 'Vulnerability Name', 'Risk', 'Severity', 'CVE/CWE ID', 'URLs','Scanner']]

        for vl in allresults:

            count = 0
            for vuln in sorted(vl, key=lambda x: x['severity'], reverse=True):
                count += 1

                name = vuln['name']

                risk = vuln['risk']
                severity = vuln['severity']
                cve_id = vuln.get('cweid') or vuln.get('cve_id', '')
                cve_id=cve_id[:13]
                urls = vuln.get('url', [])
                description = vuln['description']


                solution = vuln['solution']
                reported_by = vuln['reported_by']
                description = 'N/A'
                solution = 'N/A'
                urls = f'({len(urls)} URLs) {urls}' if urls else ''
                urls=urls[:50]
                name = name[:70]
                if vl is allresults[0]:
                   zscan_report.append([count, name, risk, severity, cve_id, urls,  reported_by])

                elif vl is allresults[1]:
                   nscan_report.append([count, name, risk, severity, cve_id, urls,  reported_by])
                elif vl is allresults[2]:
                    oscan_report.append([count, name, risk, severity, cve_id, urls,  reported_by])

        try:
            conresults = self.consolidation.consolidation(zscan_report, nscan_report, oscan_report,dup)
        except:
            print("Error: unable to consolidate")

        conresults = sorted(conresults, key=lambda x: x[3], reverse=True)
        final.extend(conresults)

        try:
            val = self.overall_risk(conresults)
        except:
            print("Error: unable to calculate overall risk score")
        table0 = SingleTable(val)
        table0.inner_row_border = True
        print(table0.table)

        table = SingleTable(final)
        table.title = 'Vuln. Alerts'
        table.inner_heading_row_border = True
        table.inner_row_border = True
        print(table.table)

        table1 = SingleTable(dup)
        table1.title = 'Duplicate vulnerabilities'
        table1.inner_heading_row_border = True
        table1.inner_row_border = True
        print(table1.table)
        return final

    def export(self, scan_name, zscan_results, nscan_results, oscan_results):

        # if not zscan_results and nscan_results and oscan_results :
        # print ("scanning result from one or more scanner is not complete")
        # return False

        results1 = list(zscan_results.values())
        results2 = list(nscan_results.values())
        results3 = list(oscan_results.values())
        allresults = [results1, results2, results3]

        zscan_report = []
        nscan_report = []
        oscan_report = []

        final = [['#', 'Vulnerability Name', 'Risk', 'Severity', 'CVE/CWE ID', 'URLs', 'Description.', 'Solution.',
                  'Scanner']]

        dup1=[['#', 'Vulnerability Name', 'Risk', 'Severity', 'CVE/CWE ID', 'URLs', 'Description.', 'Solution.', 'Scanner']]
        for vl in allresults:

            count = 0
            for vuln in sorted(vl, key=lambda x: x['severity'], reverse=True):
                count += 1

                name = vuln['name']

                risk = vuln['risk']
                severity = vuln['severity']
                cve_id = vuln.get('cweid') or vuln.get('cve_id', '')
                cve_id = cve_id[:13]
                urls = vuln.get('url', [])
                description = vuln['description']
                sol = vuln['solution']
                reported_by = vuln['reported_by']

                urls = f'({len(urls)} URLs) {urls}' if urls else ''

                if vl == allresults[0]:
                    zscan_report.append([count, name, risk, severity, cve_id, urls, description, sol, reported_by])
                elif vl == allresults[1]:
                    nscan_report.append([count, name, risk, severity, cve_id, urls, description, sol, reported_by])
                elif vl == allresults[2]:
                    oscan_report.append([count, name, risk, severity, cve_id, urls, description, sol, reported_by])
        try:
            conresults = self.consolidation.consolidation2(zscan_report, nscan_report, oscan_report,dup1)
        except:
            print("Error: unable to consolidate")

        conresults = sorted(conresults, key=lambda x: x[3],reverse=True)
        final.extend(conresults)


        val = self.overall_risk(conresults)

        asc = re.compile(r'\x1b[^m]*m')

        wb = load_workbook('Report.xlsx')
        sheet = wb.create_sheet(scan_name)

        vl1 = sheet.cell(1, 2)
        vl1.value = "Overall Risk Score"
        vl2 = sheet.cell(1, 3)
        vl2.value =  float(asc.sub('',val[0][1]))

        vl3 = sheet.cell(2, 2)
        vl3.value = "High"
        vl4 = sheet.cell(2, 3)
        vl4.value = float(asc.sub('', val[1][1]))

        vl5 = sheet.cell(3, 2)
        vl5.value = "Medium"
        vl6 = sheet.cell(3, 3)
        vl6.value = float(asc.sub('', val[2][1]))

        vl7 = sheet.cell(4, 2)
        vl7.value = "Low"
        vl8 = sheet.cell(4, 3)
        vl8.value = float(asc.sub('', val[3][1]))

        vl9 = sheet.cell(5, 2)
        vl9.value = "Total"
        vl10 = sheet.cell(5, 3)
        vl10.value = float(asc.sub('', val[4][1]))


        row = 7
        col = 1

        for n, Name, Risk, Severity, ID, URLs, Desc, Sol, Scanner in final:

           v1 = sheet.cell(row, col)
           v1.value = n
           v2 = sheet.cell(row, col + 1)
           v2.value = Name
           v3 = sheet.cell(row, col + 2)
           v3.value = Risk
           v4 = sheet.cell(row, col + 3)
           v4.value = Severity
           v5 = sheet.cell(row, col + 4)
           v5.value = ID
           v6 = sheet.cell(row, col + 5)
           v6.value = URLs
           v7 = sheet.cell(row, col + 6)
           v7.value = Desc
           v8 = sheet.cell(row, col + 7)
           v8.value = Sol
           v9 = sheet.cell(row, col + 8)
           v9.value = Scanner

           row += 1

        col = 1
        row1=row+2
        for n, Name, Risk, Severity, ID, URLs, Desc, Sol, Scanner in dup1:
            v11 = sheet.cell(row1, col)
            v11.value = n
            v21 = sheet.cell(row1, col + 1)
            v21.value = Name
            v31 = sheet.cell(row1, col + 2)
            v31.value = Risk
            v41 = sheet.cell(row1, col + 3)
            v41.value = Severity
            v51 = sheet.cell(row1, col + 4)
            v51.value = ID
            v61 = sheet.cell(row1, col + 5)
            v61.value = URLs
            v71 = sheet.cell(row1, col + 6)
            v71.value = Desc
            v81 = sheet.cell(row1, col + 7)
            v81.value = Sol
            v91 = sheet.cell(row1, col + 8)
            v91.value = Scanner

            row1 += 1

        wb.save('Report.xlsx')

        print("Sucessfully Report exported to Excel file ")


    def overall_risk(self, consresults):
        n_l = 0
        n_m = 0
        n_h = 0
        rh = 0.0
        rm = 0.0
        rl = 0.0
        for i in range(len(consresults)):
            score = consresults[i][3]

            if 0.1 <= score <= 3.9:
                n_l = (n_l + 1)
                rl = (rl + score)
            elif 4 <= score <= 6.9:
                n_m = (n_m + 1)
                rm = (rm + score)
            elif 7 <= score <= 10:
                n_h = (n_h + 1)
                rh = (rh + score)
            else:
                continue

        R = ((rl * 0.2) + (rm * 0.545) + (rh * 0.85))
        T = (n_l + n_m + n_h)
        return [[colored('Overall risk score', 'green', ), colored(R, 'green', )],
                [colored('High', 'red', ), colored(n_h, 'red', )],
                [colored("medium", 'blue'), colored(n_m, 'blue', )],
                [colored("Low", 'yellow', ), colored(n_l, 'yellow', )],
                [colored("Total Vulnerabilities ", 'magenta', ), colored(T, 'magenta')]]
